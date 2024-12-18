import argparse
import copy
import os
import tempfile
import time
import traceback
import uuid
from datetime import datetime, timedelta
from io import StringIO
from threading import Thread

import dotenv
import identity.flask
import pandas as pd
from azure.data.tables import TableServiceClient
from flask import (
    Flask,
    copy_current_request_context,
    jsonify,
    render_template,
    request,
    send_file,
    session,
)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from werkzeug.middleware.proxy_fix import ProxyFix

from flask_session import Session
from viewer import Searching, app_config

dotenv.load_dotenv(override=True)

__version__ = "1.1.0-beta"

app = Flask(__name__)
app.config.from_object(app_config)
assert app.config["REDIRECT_PATH"] != "/", "REDIRECT_PATH must not be /"
Session(app)

app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

auth = identity.flask.Auth(
    app,
    authority=app.config["AUTHORITY"],
    client_id=app.config["CLIENT_ID"],
    client_credential=app.config["CLIENT_SECRET"],
    redirect_uri=app.config["REDIRECT_URI"],
)


connection_string = f"AccountName={os.getenv('AZURE_STORAGE_ACCOUNT_NAME')};AccountKey={os.getenv('AZURE_STORAGE_ACCOUNT_KEY')};EndpointSuffix=core.windows.net"
client = TableServiceClient.from_connection_string(conn_str=connection_string)
searchlogs = client.create_table_if_not_exists(table_name="searchlogs")
resultslogs = client.create_table_if_not_exists(table_name="resultslogs")
errorlogs = client.create_table_if_not_exists(table_name="errorlogs")


searcher = Searching.SearchEngine(os.environ["db_URI"])

data_last_updated_date = searcher.all_document_types_table.list_versions()[-1][
    "timestamp"
].strftime("%Y-%m-%d")

app.jinja_env.globals.update(
    version=__version__, data_last_updated_date=data_last_updated_date
)


def log_search(search: Searching.Search, user: str):
    if searchlogs:
        search_log = {
            "PartitionKey": user["name"],
            "RowKey": search.uuid.hex,
            "query": search.get_query(),
            "start_time": search.creation_time,
            **search.get_settings().to_dict(),
        }
        try:
            searchlogs.create_entity(entity=search_log)
        except Exception as e:
            print(f"Error logging search: {e}")
    else:
        print("Error table does not exist")


def log_search_results(results: Searching.SearchResult, user: str):
    if resultslogs:
        results_log = {
            "PartitionKey": user["name"],
            "RowKey": results.search.uuid.hex,
            "duration": results.duration,
            "summary": results.get_summary(),
            "search_results": (
                results.get_context_cleaned()
                .head(100)
                .drop(columns=["document"])
                .to_json()
            ),
            "num_results": results.get_context().shape[0],
        }
        try:
            resultslogs.create_entity(entity=results_log)
        except Exception as e:
            print(f"Error logging results: {e}")
    else:
        print("Error table does not exist")


def log_search_error(e, search: Searching.Search, user: str):
    if errorlogs:
        error_log = {
            "PartitionKey": user["name"],
            "RowKey": search.uuid.hex,
            "error": repr(e),
        }
        try:
            errorlogs.create_entity(entity=error_log)
        except Exception as e:
            print(f"Error logging search error: {e}")
    else:
        print("Error table does not exist")


@app.route("/")
@auth.login_required
def index(*, context):
    return render_template(
        "index.html",
        user=context["user"],
    )


@app.route("/feedback")
@auth.login_required
def feedback(*, context):
    return render_template(
        "feedback_form.html",
        user=context["user"],
        feedback_form_loaded=True,
    )


tasks = {}


@app.before_request
def setup_task_deleter():
    app.before_request_funcs[None].remove(setup_task_deleter)
    Thread(target=delete_old_tasks, daemon=True).start()


def delete_old_tasks():
    print("Starting delete old task loop")
    while True:
        now = datetime.now()
        one_day_ago = now - timedelta(minutes=11)
        tasks_to_delete = [
            task_id
            for task_id, task in tasks.items()
            if task.creation_time < one_day_ago
        ]
        for task_id in tasks_to_delete:
            del tasks[task_id]
        time.sleep(600)  # Sleep for one hour


class Task:
    def __init__(self):
        self.status = "in progress"
        self.result = None
        self.status_desc = None
        self.creation_time = datetime.now()

    def update(self, status, desc=None, result=None):
        if self.status == "completed" and result is None:
            raise ValueError("result must be provided if status is completed")
        self.status = status
        self.result = result
        self.status_desc = desc

    def get_status(self):
        return (self.status, self.status_desc)

    def get_result(self):
        return self.result


def create_task() -> str:
    task_id = str(uuid.uuid4())
    tasks[task_id] = Task()
    return task_id


@app.route("/task-status/<task_id>", methods=["GET"])
@auth.login_required
def task_status(task_id, *, context):
    task = tasks.get(task_id)
    result = task.get_result() if task else None
    status, status_desc = task.get_status() if task else ("not found", None)
    print(f"Task status: '{status}'")
    jsonified = jsonify(
        {
            "task_id": task_id,
            "status": status,
            "status_desc": status_desc,
            "result": result,
        }
    )
    if status == "completed":
        session["search_results"] = result
        del tasks[task_id]
    return jsonified


@app.route("/search", methods=["POST"])
@auth.login_required
def search(*, context):
    form_data = request.form

    task_id = create_task()
    task_thread = Thread(
        target=copy_current_request_context(search_reports),
        args=(task_id, form_data, context),
    )
    task_thread.start()
    return jsonify({"task_id": task_id}), 202


def search_reports(task_id, form_data, context):
    task = tasks.get(task_id)
    try:
        search = Searching.Search.from_form(form_data)
        log_search(search, context["user"])
        results = searcher.search(search)
        while True:
            try:
                update = next(results)
                task.update(status="in progress", desc=update)
            except StopIteration as e:
                results = e.value
                break
        formatted_results = format_search_results(results)
        task.update("completed", None, formatted_results)
        log_search_results(results, context["user"])
    except Exception as e:
        print("Error while doing search:\n" + "".join(traceback.format_exception(e)))
        log_search_error(e, search, context["user"])
        task.update("failed", None, repr(e))
        return


def get_updated_relevance_search(search, new_relevance):
    search.settings.relevanceCutoff = new_relevance
    return search.to_url_params()


def format_search_results(results: Searching.SearchResult):
    context_df = results.get_context_cleaned()

    context_df["report_id"] = context_df[["report_id", "url"]].apply(
        lambda x: f'<a href="{x["url"]}" target="_blank">{x["report_id"]}</a>', axis=1
    )

    context_df = context_df.drop(columns=["url"])

    context_df["relevance"] = context_df.apply(
        lambda x: f"""<a href="/?{get_updated_relevance_search(copy.deepcopy(results.search), x['relevance'])}">{x['relevance']}</a>""",
        axis=1,
    )

    html_table = context_df.to_html(
        classes="table table-bordered table-hover align-middle",
        table_id="dataTable",
        justify="center",
        index=False,
        escape=False,
    )

    document_type_pie_chart = results.get_document_type_pie_chart().to_json()

    mode_pie_chart = results.get_mode_pie_chart().to_json()

    year_hist = results.get_year_histogram().to_json()

    most_common_event_types = results.get_most_common_event_types().to_json()

    agency_distribution = results.get_agency_pie_chart().to_json()

    print(f"Formatted results {len(context_df)}")

    return {
        "html_table": html_table,
        "results_summary_info": {
            "document_type_pie_chart": document_type_pie_chart,
            "mode_pie_chart": mode_pie_chart,
            "year_histogram": year_hist,
            "most_common_event_types": most_common_event_types,
            "agency_pie_chart": agency_distribution,
            "duration": results.get_search_duration_str(),
            "quick_search_summary": f"There are {context_df.shape[0]} results across {context_df['report_id'].unique().shape[0]} reports",
            "num_results": context_df.shape[0],
        },
        "summary": results.get_summary(),
        "settings": results.search.get_settings().to_dict(),
        "start_time": results.search.get_start_time(),
        "query": results.search.get_query(),
    }


def send_csv_file(df: pd.DataFrame, name: str):
    temp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False)

    # Write the CSV data to the file
    df.to_csv(temp.name, index=False)

    # Send the file
    return send_file(temp.name, as_attachment=True, download_name=name)


@app.route("/get_results_as_csv", methods=["POST"])
@auth.login_required
def get_results_as_csv(*, context):
    if session.get("search_results") is None:
        return jsonify({"error": "No results found"}), 404

    search_results = session.get("search_results")

    # Create a temporary file
    temp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temp.close()  # Close the file so it can be opened by openpyxl

    # Create a new Workbook
    wb = Workbook()

    # Write summary information to the first sheet
    summary_info = search_results["results_summary_info"]
    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    summary_sheet["A1"] = "Search Query:"
    summary_sheet["D1"] = "Redo search:"
    url_params_for_search = Searching.Search(
        search_results["query"],
        settings=Searching.SearchSettings.from_dict(search_results["settings"]),
    ).to_url_params()
    summary_sheet["E1"].hyperlink = f"""{request.url_root}?{url_params_for_search}"""
    summary_sheet["A2"] = search_results["query"]

    summary_sheet["A4"] = "Start Time:"
    summary_sheet["A5"] = search_results["start_time"]

    summary_sheet["A7"] = "Settings:"
    summary_sheet["A8"] = repr(search_results["settings"])

    summary_sheet["A10"] = "Search Duration:"
    summary_sheet["A11"] = summary_info["duration"]

    summary_sheet["A13"] = "Number of Results:"
    summary_sheet["A14"] = summary_info["num_results"]

    summary_sheet["A16"] = "Summary:"
    summary_sheet["A17"] = search_results["summary"]

    # Create a second sheet and write the search results using pandas
    results_df = pd.read_html(StringIO(search_results["html_table"]), flavor="lxml")[0]

    results_sheet = wb.create_sheet(title="Search Results")

    for r_idx, row in enumerate(
        dataframe_to_rows(results_df, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            results_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook to the temporary file
    wb.save(temp.name)

    # Send the file as a response
    search_query = search_results["query"]
    search_start_time = search_results["start_time"]
    download_name = f"{search_query}-{search_start_time}.xlsx"

    return send_file(temp.name, as_attachment=True, download_name=download_name)


def run():
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", action="store_true", help="Enable debug mode")
    args = parser.parse_args()

    app.run(port=5000, host="localhost", debug=args.debug)


if __name__ == "__main__":
    app.run()
